from aiogram import Router, F
from aiogram.types import CallbackQuery, BufferedInputFile
from app.database.models import async_session, Ticket, Status, Company, Category, Ticket_category, Module
from openpyxl import Workbook
from io import BytesIO
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from datetime import timezone
from zoneinfo import ZoneInfo
import app.admin_keyboard as admin_kb

router = Router()

@router.callback_query(F.data == 'export_report')
async def ticket(callback:CallbackQuery):
    await callback.answer('')
    await callback.message.edit_text('Выберите тип отчета', reply_markup=admin_kb.choice_report()) 

@router.callback_query(F.data == 'report_status')
async def ticket(callback:CallbackQuery):
    await callback.answer('')
    await callback.message.edit_text('Выберите статус', reply_markup=admin_kb.status_choice_report()) 

@router.callback_query(F.data == 'report_company')
async def ticket(callback:CallbackQuery):
    await callback.answer('') 
    await callback.message.edit_text('Выберите компанию', reply_markup=admin_kb.company_choice_report()) 

@router.callback_query(F.data == 'report_category')
async def ticket(callback:CallbackQuery):
    await callback.answer('') 
    await callback.message.edit_text('Выберите тип замечания', reply_markup=admin_kb.category_choice_report()) 


#ОБРАБОТКА ВОЗВРАТОВ
@router.callback_query(F.data == "back_fr_report")
async def back_tick(callback:CallbackQuery):
    await callback.answer('')
    await callback.message.edit_text('Что хотите сделать?', reply_markup=admin_kb.status_or_report) 

@router.callback_query(F.data == "back_fr_reportst")
async def back_tickst(callback:CallbackQuery):
    await callback.answer('')
    await callback.message.edit_text('Выберите тип отчета', reply_markup=admin_kb.choice_report()) 

@router.callback_query(F.data == "back_fr_reportcm")
async def back_tickcm(callback:CallbackQuery):
    await callback.answer('')
    await callback.message.edit_text('Выберите тип отчета', reply_markup=admin_kb.choice_report()) 

@router.callback_query(F.data == "back_fr_reportct")
async def back_tickct(callback:CallbackQuery):
    await callback.answer('')
    await callback.message.edit_text('Выберите тип отчета', reply_markup=admin_kb.choice_report()) 


#ВЫГРУЗКА ОТЧЕТОВ ПО ЗАЯВКАМ
#выгрузка отчет по всем заявкам
@router.callback_query(F.data == "export_full_report")
async def export_full_report_handler(callback: CallbackQuery):
    await callback.answer()

    async with async_session() as session:
        result = await session.execute(
            select(Ticket)
            .options(selectinload(Ticket.company), selectinload(Ticket.status), selectinload(Ticket.user), selectinload(Ticket.module))
            .order_by(Ticket.ticket_id)
        )
        tickets = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"

    headers = ["Номер заявки", "Имя пользователя", "Компания", "Модуль", "Описание", "Статус", "Дата создания", "Комментарий админа"]
    ws.append(headers)

    for t in tickets:
        if t.data_created:
            local_time = t.data_created.replace(tzinfo=timezone.utc).astimezone(ZoneInfo("Europe/Moscow"))
            created_str = local_time.strftime("%Y-%m-%d %H:%M:%S")
        else:
            created_str = "–"
        ws.append([
            t.ticket_id,
            t.user.name if t.user else "-",
            t.company.company_name if t.company else "–",
            t.module.name if t.module else "–",
            t.description,
            t.status.name if t.status else "–",
            created_str,
            t.comment_adm
        ])

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    input_file=BufferedInputFile(file=file_stream.read(), filename="Все заявки.xlsx")
    await callback.message.answer_document(
        document=input_file,
        caption="Отчет по всем заявкам"
    )

#выгрузка отчет по новым заявкам
@router.callback_query(F.data == "export_new_report")
async def export_new_report_handler(callback: CallbackQuery):
    await callback.answer()

    async with async_session() as session:
        result = await session.execute(
            select(Ticket)
            .join(Ticket.status)
            .options(selectinload(Ticket.company), selectinload(Ticket.status), selectinload(Ticket.user), selectinload(Ticket.module))
            .filter(Status.name == "Новая")
            .order_by(Ticket.ticket_id)
        )
        tickets = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"

    headers = ["Номер заявки", "Имя пользователя", "Компания", "Модуль", "Описание", "Статус", "Дата создания"]
    ws.append(headers)

    for t in tickets:
        if t.data_created:
            local_time = t.data_created.replace(tzinfo=timezone.utc).astimezone(ZoneInfo("Europe/Moscow"))
            created_str = local_time.strftime("%Y-%m-%d %H:%M:%S")
        else:
            created_str = "–"
        ws.append([
            t.ticket_id,
            t.user.name if t.user else "-",
            t.company.company_name if t.company else "–",
            t.module.name if t.module else "–",
            t.description,
            t.status.name if t.status else "–",
            created_str
        ])

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    input_file=BufferedInputFile(file=file_stream.read(), filename="Все новые заявки.xlsx")
    await callback.message.answer_document(
        document=input_file,
        caption="Отчет по новым заявкам"
    )

#выгрузка отчет по заявкам в работе
@router.callback_query(F.data == "export_work_report")
async def export_work_report_handler(callback: CallbackQuery):
    await callback.answer()

    async with async_session() as session:
        result = await session.execute(
            select(Ticket)
            .join(Ticket.status)
            .options(selectinload(Ticket.company), selectinload(Ticket.status), selectinload(Ticket.user), selectinload(Ticket.module))
            .filter(Status.name == "В работе")
            .order_by(Ticket.ticket_id)
        )
        tickets = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"

    headers = ["Номер заявки", "Имя пользователя", "Компания", "Модуль", "Описание", "Статус", "Дата создания", "Комментарий админа"]
    ws.append(headers)

    for t in tickets:
        if t.data_created:
            local_time = t.data_created.replace(tzinfo=timezone.utc).astimezone(ZoneInfo("Europe/Moscow"))
            created_str = local_time.strftime("%Y-%m-%d %H:%M:%S")
        else:
            created_str = "–"
        ws.append([
            t.ticket_id,
            t.user.name if t.user else "-",
            t.company.company_name if t.company else "–",
            t.module.name if t.module else "–",
            t.description,
            t.status.name if t.status else "–",
            created_str,
            t.comment_adm
        ])

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    input_file=BufferedInputFile(file=file_stream.read(), filename="Все заявки в работе.xlsx")
    await callback.message.answer_document(
        document=input_file,
        caption="Отчет по заявкам в работе"
    )

#выгрузка отчет по заявкам "Завершена"
@router.callback_query(F.data == "export_completed_report")
async def export_completed_report_handler(callback: CallbackQuery):
    await callback.answer()

    async with async_session() as session:
        result = await session.execute(
            select(Ticket)
            .join(Ticket.status)
            .options(selectinload(Ticket.company), selectinload(Ticket.status), selectinload(Ticket.user), selectinload(Ticket.module))
            .filter(Status.name == "Завершена")
            .order_by(Ticket.ticket_id)
        )
        tickets = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"

    headers = ["Номер заявки", "Имя пользователя", "Компания", "Модуль", "Описание", "Статус", "Дата создания", "Комментарий от админа"]
    ws.append(headers)

    for t in tickets:
        if t.data_created:
            local_time = t.data_created.replace(tzinfo=timezone.utc).astimezone(ZoneInfo("Europe/Moscow"))
            created_str = local_time.strftime("%Y-%m-%d %H:%M:%S")
        else:
            created_str = "–"
        ws.append([
            t.ticket_id,
            t.user.name if t.user else "-",
            t.company.company_name if t.company else "–",
            t.module.name if t.module else "–",
            t.description,
            t.status.name if t.status else "–",
            created_str,
            t.comment_adm
        ])

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    input_file=BufferedInputFile(file=file_stream.read(), filename="Все завершенные заявки.xlsx")
    await callback.message.answer_document(
        document=input_file,
        caption="Отчет по завершенным заявкам"
    )

#выгрузка отчет по Новатексу
@router.callback_query(F.data == "novateks_report")
async def export_novateks_report_handler(callback: CallbackQuery):
    await callback.answer()

    async with async_session() as session:
        result = await session.execute(
            select(Ticket)
            .join(Ticket.company)
            .options(selectinload(Ticket.company), selectinload(Ticket.status), selectinload(Ticket.user), selectinload(Ticket.module))
            .filter(Company.company_name == "Новатекс")
            .order_by(Ticket.ticket_id)
        )
        tickets = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"

    headers = ["Номер заявки", "Имя пользователя", "Компания", "Модуль", "Описание", "Статус", "Дата создания", "Комментарий от админа"]
    ws.append(headers)

    for t in tickets:
        if t.data_created:
            local_time = t.data_created.replace(tzinfo=timezone.utc).astimezone(ZoneInfo("Europe/Moscow"))
            created_str = local_time.strftime("%Y-%m-%d %H:%M:%S")
        else:
            created_str = "–"
        ws.append([
            t.ticket_id,
            t.user.name if t.user else "-",
            t.company.company_name if t.company else "–",
            t.module.name if t.module else "–",
            t.description,
            t.status.name if t.status else "–",
            created_str,
            t.comment_adm
        ])

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    input_file=BufferedInputFile(file=file_stream.read(), filename="Все заявки Новатекс.xlsx")
    await callback.message.answer_document(
        document=input_file,
        caption="Отчет по заявкам Новатекс"
    )

#выгрузка отчет по ЗДИ
@router.callback_query(F.data == "zdi_report")
async def export_zdi_report_handler(callback: CallbackQuery):
    await callback.answer()

    async with async_session() as session:
        result = await session.execute(
            select(Ticket)
            .join(Ticket.company)
            .options(selectinload(Ticket.company), selectinload(Ticket.status), selectinload(Ticket.user), selectinload(Ticket.module))
            .filter(Company.company_name == "ЗДИ")
            .order_by(Ticket.ticket_id)
        )
        tickets = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"

    headers = ["Номер заявки", "Имя пользователя", "Компания", "Модуль", "Описание", "Статус", "Дата создания", "Комментарий админа"]
    ws.append(headers)

    for t in tickets:
        if t.data_created:
            local_time = t.data_created.replace(tzinfo=timezone.utc).astimezone(ZoneInfo("Europe/Moscow"))
            created_str = local_time.strftime("%Y-%m-%d %H:%M:%S")
        else:
            created_str = "–"
        ws.append([
            t.ticket_id,
            t.user.name if t.user else "-",
            t.company.company_name if t.company else "–",
            t.module.name if t.module else "–",
            t.description,
            t.status.name if t.status else "–",
            created_str,
            t.comment_adm
        ])

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    input_file=BufferedInputFile(file=file_stream.read(), filename="Все заявки ЗДИ.xlsx")
    await callback.message.answer_document(
        document=input_file,
        caption="Отчет по заявкам ЗДИ"
    )

#выгрузка отчет по Интерскол
@router.callback_query(F.data == "interskol_report")
async def export_interskol_report_handler(callback: CallbackQuery):
    await callback.answer()

    async with async_session() as session:
        result = await session.execute(
            select(Ticket)
            .join(Ticket.company)
            .options(selectinload(Ticket.company), selectinload(Ticket.status), selectinload(Ticket.user), selectinload(Ticket.module))
            .filter(Company.company_name == "Интерскол")
            .order_by(Ticket.ticket_id)
        )
        tickets = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"

    headers = ["Номер заявки", "Имя пользователя", "Компания", "Модуль", "Описание", "Статус", "Дата создания", "Комментарий админа"]
    ws.append(headers)

    for t in tickets:
        if t.data_created:
            local_time = t.data_created.replace(tzinfo=timezone.utc).astimezone(ZoneInfo("Europe/Moscow"))
            created_str = local_time.strftime("%Y-%m-%d %H:%M:%S")
        else:
            created_str = "–"
        ws.append([
            t.ticket_id,
            t.user.name if t.user else "-",
            t.company.company_name if t.company else "–",
            t.module.name if t.module else "–",
            t.description,
            t.status.name if t.status else "–",
            created_str,
            t.comment_adm
        ])

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    input_file=BufferedInputFile(file=file_stream.read(), filename="Все заявки Интерскол.xlsx")
    await callback.message.answer_document(
        document=input_file,
        caption="Отчет по заявкам Интерскол"
    )
   
#выгрузка отчет по Сова
@router.callback_query(F.data == "sova_report")
async def export_sova_report_handler(callback: CallbackQuery):
    await callback.answer()

    async with async_session() as session:
        result = await session.execute(
            select(Ticket)
            .join(Ticket.company)
            .options(selectinload(Ticket.company), selectinload(Ticket.status), selectinload(Ticket.user), selectinload(Ticket.module))
            .filter(Company.company_name == "Сова Моторс Алабуга")
            .order_by(Ticket.ticket_id)
        )
        tickets = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"

    headers = ["Номер заявки", "Имя пользователя", "Компания", "Модуль", "Описание", "Статус", "Дата создания", "Комментарий админа"]
    ws.append(headers)

    for t in tickets:
        if t.data_created:
            local_time = t.data_created.replace(tzinfo=timezone.utc).astimezone(ZoneInfo("Europe/Moscow"))
            created_str = local_time.strftime("%Y-%m-%d %H:%M:%S")
        else:
            created_str = "–"
        ws.append([
            t.ticket_id,
            t.user.name if t.user else "-",
            t.company.company_name if t.company else "–",
            t.module.name if t.module else "–",
            t.description,
            t.status.name if t.status else "–",
            created_str,
            t.comment_adm
        ])

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    input_file=BufferedInputFile(file=file_stream.read(), filename="Все заявки Сова Моторс Алабуга.xlsx")
    await callback.message.answer_document(
        document=input_file,
        caption="Отчет по заявкам Сова Моторс Алабуга"
    )

#выгрузка отчет по Алтрест
@router.callback_query(F.data == "altrest_report")
async def export_altrest_report_handler(callback: CallbackQuery):
    await callback.answer()

    async with async_session() as session:
        result = await session.execute(
            select(Ticket)
            .join(Ticket.company)
            .options(selectinload(Ticket.company), selectinload(Ticket.status), selectinload(Ticket.user))
            .filter(Company.company_name == "Алтрест")
            .order_by(Ticket.ticket_id)
        )
        tickets = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"

    headers = ["Номер заявки", "Имя пользователя", "Компания", "Описание", "Статус", "Дата создания", "Комментарий админа"]
    ws.append(headers)

    for t in tickets:
        if t.data_created:
            local_time = t.data_created.replace(tzinfo=timezone.utc).astimezone(ZoneInfo("Europe/Moscow"))
            created_str = local_time.strftime("%Y-%m-%d %H:%M:%S")
        else:
            created_str = "–"
        ws.append([
            t.ticket_id,
            t.user.name if t.user else "-",
            t.company.company_name if t.company else "–",
            t.description,
            t.status.name if t.status else "–",
            created_str,
            t.comment_adm
        ])

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    input_file=BufferedInputFile(file=file_stream.read(), filename="Все заявки Алтрест.xlsx")
    await callback.message.answer_document(
        document=input_file,
        caption="Отчет по заявкам Алтрест"
    )

#выгрузка отчет по Тексфлор
@router.callback_query(F.data == "teksflor_report")
async def export_teksflor_report_handler(callback: CallbackQuery):
    await callback.answer()

    async with async_session() as session:
        result = await session.execute(
            select(Ticket)
            .join(Ticket.company)
            .options(selectinload(Ticket.company), selectinload(Ticket.status), selectinload(Ticket.user))
            .filter(Company.company_name == "Тексфлор")
            .order_by(Ticket.ticket_id)
        )
        tickets = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"

    headers = ["Номер заявки", "Имя пользователя", "Компания", "Описание", "Статус", "Дата создания", "Комментарий админа"]
    ws.append(headers)

    for t in tickets:
        if t.data_created:
            local_time = t.data_created.replace(tzinfo=timezone.utc).astimezone(ZoneInfo("Europe/Moscow"))
            created_str = local_time.strftime("%Y-%m-%d %H:%M:%S")
        else:
            created_str = "–"
        ws.append([
            t.ticket_id,
            t.user.name if t.user else "-",
            t.company.company_name if t.company else "–",
            t.description,
            t.status.name if t.status else "–",
            created_str,
            t.comment_adm
        ])

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    input_file=BufferedInputFile(file=file_stream.read(), filename="Все заявки Тексфлор.xlsx")
    await callback.message.answer_document(
        document=input_file,
        caption="Отчет по заявкам Тексфлор"
    )

#выгрузка заявок по категории мелкосрочный ремонт
@router.callback_query(F.data == "remont_report")
async def export_remont_report_handler(callback: CallbackQuery):
    await callback.answer()
    category_name = "Мелкосрочный ремонт"
    async with async_session() as session:
        result = await session.execute(
            select(Ticket)
            .join(Ticket.categories)
            .options(
            selectinload(Ticket.user),
            selectinload(Ticket.company),
            selectinload(Ticket.status),
            selectinload(Ticket.categories),
            selectinload(Ticket.module)
            )
            .filter(Category.name == category_name)
            .order_by(Ticket.ticket_id)
        )
        tickets = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"

    headers = ["Номер заявки", "Имя пользователя", "Компания", "Модуль", "Описание", "Статус", "Дата создания", "Категория", "Комментарий админа"]
    ws.append(headers)

    for t in tickets:
        created_str = t.data_created.astimezone(ZoneInfo("Europe/Moscow")).strftime("%Y-%m-%d %H:%M:%S") if t.data_created else "-"
        categories_str = ", ".join([c.name for c in t.categories])
        ws.append([
            t.ticket_id,
            t.user.name if t.user else "-",
            t.company.company_name if t.company else "-",
            t.module.name if t.module else "–",
            t.description,
            t.status.name if t.status else "-",
            created_str,
            categories_str,
            t.comment_adm
        ])

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    input_file=BufferedInputFile(file=file_stream.read(), filename="Все заявки по Мелкосрочному ремонту.xlsx")
    await callback.message.answer_document(
        document=input_file,
        caption="Отчет по заявкам 'Мелкосрочный ремонт'"
    )

#выгрузка заявок по категории Прилегающая территория
@router.callback_query(F.data == "territory_report")
async def export_territory_report_handler(callback: CallbackQuery):
    await callback.answer()
    category_name = "Прилегающая территория"
    async with async_session() as session:
        result = await session.execute(
            select(Ticket)
            .join(Ticket.categories)
            .options(
            selectinload(Ticket.user),
            selectinload(Ticket.company),
            selectinload(Ticket.status),
            selectinload(Ticket.categories),
            selectinload(Ticket.module)
            )
            .filter(Category.name == category_name)
            .order_by(Ticket.ticket_id)
        )
        tickets = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"

    headers = ["Номер заявки", "Имя пользователя", "Компания", "Модуль", "Описание", "Статус", "Дата создания", "Категория", "Комментарий админа"]
    ws.append(headers)

    for t in tickets:
        created_str = t.data_created.astimezone(ZoneInfo("Europe/Moscow")).strftime("%Y-%m-%d %H:%M:%S") if t.data_created else "-"
        categories_str = ", ".join([c.name for c in t.categories])
        ws.append([
            t.ticket_id,
            t.user.name if t.user else "-",
            t.company.company_name if t.company else "-",
            t.module.name if t.module else "–",
            t.description,
            t.status.name if t.status else "-",
            created_str,
            categories_str,
            t.comment_adm
        ])

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    input_file=BufferedInputFile(file=file_stream.read(), filename="Все заявки по Прилегающей территории.xlsx")
    await callback.message.answer_document(
        document=input_file,
        caption="Отчет по заявкам 'Прилегающая территория'"
    )

#выгрузка заявок по категории Корректировка систем теплоснабжения
@router.callback_query(F.data == "correct_warm_report")
async def export_correct_warm_report_handler(callback: CallbackQuery):
    await callback.answer()
    category_name = "Корректировка систем теплоснабжения"
    async with async_session() as session:
        result = await session.execute(
            select(Ticket)
            .join(Ticket.categories)
            .options(
            selectinload(Ticket.user),
            selectinload(Ticket.company),
            selectinload(Ticket.status),
            selectinload(Ticket.categories),
            selectinload(Ticket.module)
            )
            .filter(Category.name == category_name)
            .order_by(Ticket.ticket_id)
        )
        tickets = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"

    headers = ["Номер заявки", "Имя пользователя", "Компания", "Модуль", "Описание", "Статус", "Дата создания", "Категория", "Комментарий админа"]
    ws.append(headers)

    for t in tickets:
        created_str = t.data_created.astimezone(ZoneInfo("Europe/Moscow")).strftime("%Y-%m-%d %H:%M:%S") if t.data_created else "-"
        categories_str = ", ".join([c.name for c in t.categories])
        ws.append([
            t.ticket_id,
            t.user.name if t.user else "-",
            t.company.company_name if t.company else "-",
            t.module.name if t.module else "–",
            t.description,
            t.status.name if t.status else "-",
            created_str,
            categories_str,
            t.comment_adm
        ])

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    input_file=BufferedInputFile(file=file_stream.read(), filename="Все заявки по Корректировке систем теплоснабжения.xlsx")
    await callback.message.answer_document(
        document=input_file,
        caption="Отчет по заявкам 'Корректировка систем теплоснабжения'"
    )

#выгрузка заявок по категории Водоснабжение/водоотведение
@router.callback_query(F.data == "water_report")
async def export_water_report_handler(callback: CallbackQuery):
    await callback.answer()
    category_name = "Водоснабжение/водоотведение"
    async with async_session() as session:
        result = await session.execute(
            select(Ticket)
            .join(Ticket.categories)
            .options(
            selectinload(Ticket.user),
            selectinload(Ticket.company),
            selectinload(Ticket.status),
            selectinload(Ticket.categories),
            selectinload(Ticket.module)
            )
            .filter(Category.name == category_name)
            .order_by(Ticket.ticket_id)
        )
        tickets = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"

    headers = ["Номер заявки", "Имя пользователя", "Компания", "Модуль", "Описание", "Статус", "Дата создания", "Категория", "Комментарий админа"]
    ws.append(headers)

    for t in tickets:
        created_str = t.data_created.astimezone(ZoneInfo("Europe/Moscow")).strftime("%Y-%m-%d %H:%M:%S") if t.data_created else "-"
        categories_str = ", ".join([c.name for c in t.categories])
        ws.append([
            t.ticket_id,
            t.user.name if t.user else "-",
            t.company.company_name if t.company else "-",
            t.module.name if t.module else "–",
            t.description,
            t.status.name if t.status else "-",
            created_str,
            categories_str,
            t.comment_adm
        ])

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    input_file=BufferedInputFile(file=file_stream.read(), filename="Все заявки по Водоснабжению, водоотведению.xlsx")
    await callback.message.answer_document(
        document=input_file,
        caption="Отчет по заявкам 'Водоснабжение/водоотведение'"
    )

#выгрузка заявок по категории Электроснабжение
@router.callback_query(F.data == "power_report")
async def export_power_report_handler(callback: CallbackQuery):
    await callback.answer()
    category_name = "Электроснабжение"
    async with async_session() as session:
        result = await session.execute(
            select(Ticket)
            .join(Ticket.categories)
            .options(
            selectinload(Ticket.user),
            selectinload(Ticket.company),
            selectinload(Ticket.status),
            selectinload(Ticket.categories),
            selectinload(Ticket.module)
            )
            .filter(Category.name == category_name)
            .order_by(Ticket.ticket_id)
        )
        tickets = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"

    headers = ["Номер заявки", "Имя пользователя", "Компания", "Модуль", "Описание", "Статус", "Дата создания", "Категория", "Комментарий админа"]
    ws.append(headers)

    for t in tickets:
        created_str = t.data_created.astimezone(ZoneInfo("Europe/Moscow")).strftime("%Y-%m-%d %H:%M:%S") if t.data_created else "-"
        categories_str = ", ".join([c.name for c in t.categories])
        ws.append([
            t.ticket_id,
            t.user.name if t.user else "-",
            t.company.company_name if t.company else "-",
            t.module.name if t.module else "–",
            t.description,
            t.status.name if t.status else "-",
            created_str,
            categories_str,
            t.comment_adm
        ])

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    input_file=BufferedInputFile(file=file_stream.read(), filename="Все заявки по Электроснабжению.xlsx")
    await callback.message.answer_document(
        document=input_file,
        caption="Отчет по заявкам 'Электроснабжение'"
    )

#выгрузка заявок по категории Прочее
@router.callback_query(F.data == "another_report")
async def export_another_report_handler(callback: CallbackQuery):
    await callback.answer()
    category_name = "Прочее"
    async with async_session() as session:
        result = await session.execute(
            select(Ticket)
            .join(Ticket.categories)
            .options(
            selectinload(Ticket.user),
            selectinload(Ticket.company),
            selectinload(Ticket.status),
            selectinload(Ticket.categories),
            selectinload(Ticket.module)
            )
            .filter(Category.name == category_name)
            .order_by(Ticket.ticket_id)
        )
        tickets = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"

    headers = ["Номер заявки", "Имя пользователя", "Компания", "Модуль", "Описание", "Статус", "Дата создания", "Категория", "Комментарий админа"]
    ws.append(headers)

    for t in tickets:
        created_str = t.data_created.astimezone(ZoneInfo("Europe/Moscow")).strftime("%Y-%m-%d %H:%M:%S") if t.data_created else "-"
        categories_str = ", ".join([c.name for c in t.categories])
        ws.append([
            t.ticket_id,
            t.user.name if t.user else "-",
            t.company.company_name if t.company else "-",
            t.module.name if t.module else "–",
            t.description,
            t.status.name if t.status else "-",
            created_str,
            categories_str,
            t.comment_adm
        ])

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    input_file=BufferedInputFile(file=file_stream.read(), filename="Все прочие заявки.xlsx")
    await callback.message.answer_document(
        document=input_file,
        caption="Отчет по заявкам 'Прочее'"
    )